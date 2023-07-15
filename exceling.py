import pickle
import random

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Fill, PatternFill, Side, Border

if __name__ == '__main__':

    lista_cadouri = [
        "Ceva de squash / sala / sport",
        "O carte",
        "Ceva cu Facultatea / Universitatea / ASMI",
        "Un tricou",
        "Ciorapi / Chiloți",
        "Ochelari de soare",
        "Ceva de pus pe cap (șapcă, etc.)",
        "Ceva Bucureștean (sau anti-provincial)",
        "Ceva la caterincă (eventual inutil)",
        "Un boardgame",
        "Ceva de îmbrăcat (nu șapcă, tricou, chiloți sau ciorapi)",
        "Vinil",
        "Joc (PC / PS4 / Xbox)",
        "Bijuterie (lanț, inel, etc.)",
        "Un cadou pe care îl mai are deja odată",
    ]

    with open('extragere.pkl', 'rb') as fp:
        extragere = pickle.load(fp)
        print('Person dictionary')
        print(extragere)

    wb = Workbook()

    for key, value in extragere.items():
        wb.create_sheet("Cardul " + str(key))

        ws1 = wb["Cardul " + str(key)]

        cell_list_odd = ['A1', 'A3', 'B1', 'B3', 'C1', 'C3', ]
        cell_list_even = ['A2', 'A4', 'B2', 'B4', 'C2', 'C4']
        cell_list = cell_list_even + cell_list_odd

        random.shuffle(value)

        for index, cell in enumerate(cell_list):
            ws1[cell] = value[index]
            ws1[cell].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            ws1[cell].border = Border(left=Side(border_style='thin', color='00666699'),
                                      right=Side(border_style='thin', color='00666699'),
                                      top=Side(border_style='thin', color='00666699')),
            if cell in cell_list_odd:
                ws1[cell].fill = PatternFill("solid", fgColor="CCC0DA")
            else:
                ws1[cell].fill = PatternFill("solid", fgColor="E4DFEC")
            if value[index] in lista_cadouri:
                ws1[cell].font = Font(size=10, color="00FF0000", bold=True, name='Congenial')
            else:
                ws1[cell].font = Font(size=10, bold=True, name='Congenial')



        ws1.column_dimensions['A'].width = 15
        ws1.column_dimensions['B'].width = 15
        ws1.column_dimensions['C'].width = 15
        ws1.column_dimensions['D'].width = 15

        ws1.row_dimensions[1].height = 78
        ws1.row_dimensions[2].height = 78
        ws1.row_dimensions[3].height = 78
        ws1.row_dimensions[4].height = 78

    wb.save('book_eg.xlsx')